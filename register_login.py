
import os  # accessing the os functions
import check_camera
import Capture_Image
import Train_Image
import Recognize

# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *


def register():
    # globally declare wb and sheet variable

    # opening the existing excel file
    wb = load_workbook("registration.xlsx")

    # create the sheet object
    sheet = wb.active


    def excel():

    	# resize the width of columns in
    	# excel spreadsheet
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 30

    	# write given data to an excel spreadsheet
    	# at particular location
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Course"
        sheet.cell(row=1, column=3).value = "Semester"
        sheet.cell(row=1, column=4).value = "Roll Number"
        sheet.cell(row=1, column=5).value = "Contact Number"
        sheet.cell(row=1, column=6).value = "Email id"
        sheet.cell(row=1, column=7).value = "Address"
        sheet.cell(row=1, column=8).value = "User Name"
        sheet.cell(row=1, column=9).value = "Password"

    global username_field
    global password_field
    # Function to set focus (cursor)
    def focus1(event):
    	# set focus on the course_field box
    	course_field.focus_set()


    # Function to set focus
    def focus2(event):
    	# set focus on the sem_field box
    	sem_field.focus_set()


    # Function to set focus
    def focus3(event):
    	# set focus on the form_no_field box
    	form_no_field.focus_set()


    # Function to set focus
    def focus4(event):
    	# set focus on the contact_no_field box
    	contact_no_field.focus_set()


    # Function to set focus
    def focus5(event):
    	# set focus on the email_id_field box
    	email_id_field.focus_set()


    # Function to set focus
    def focus6(event):
    	# set focus on the address_field box
    	address_field.focus_set()

    # Function to set focus
    def focus7(event):
    	# set focus on the username_field box
        username_field.focus_set()

    # Function to set focus
    def focus8(event):
    	# set focus on the password_field box
        password_field.focus_set()

    # Function for clearing the
    # contents of text entry boxes
    def clear():

    	# clear the content of text entry box
        name_field.delete(0, END)
        course_field.delete(0, END)
        sem_field.delete(0, END)
        form_no_field.delete(0, END)
        contact_no_field.delete(0, END)
        email_id_field.delete(0, END)
        address_field.delete(0, END)
        username_field.delete(0, END)
        password_field.delete(0, END)


    # Function to take data from GUI
    # window and write to an excel file
    def insert():

    	# if user not fill any entry
    	# then print "empty input"
    	if (name_field.get() == "" or
            course_field.get() == "" or
            sem_field.get() == "" or
            form_no_field.get() == "" or
            contact_no_field.get() == "" or
            email_id_field.get() == "" or
            address_field.get() == "" or
            username_field.get() == "" or
            password_field.get() == ""):
            root = Tk()
            w = Label(root, text='empty input')
            w.pack()
            root.mainloop()


    	else:

    		# assigning the max row and max column
    		# value upto which data is written
    		# in an excel sheet to the variable
            current_row = sheet.max_row
            current_column = sheet.max_column

    		# get method returns current text
    		# as string which we write into
    		# excel spreadsheet at particular location
            sheet.cell(row=current_row + 1, column=1).value = name_field.get()
            sheet.cell(row=current_row + 1, column=2).value = course_field.get()
            sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
            sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
            sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
            sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
            sheet.cell(row=current_row + 1, column=7).value = address_field.get()
            sheet.cell(row=current_row + 1, column=8).value = username_field.get()
            sheet.cell(row=current_row + 1, column=9).value = password_field.get()
    		# save the file
            wb.save('registration.xlsx')

    		# set focus on the name_field box
            name_field.focus_set()

    		# call the clear() function
            clear()


    # Driver code
    if __name__ == "__main__":

    	# create a GUI window
        root = Tk()

    	# set the background colour of GUI window
        root.configure(background='white')

    	# set the title of GUI window
        root.title("registration form")

    	# set the configuration of GUI window
        root.geometry("900x700")

        excel()

    	# create a Form label
        heading = Label(root, text="SIGN UP", bg="deepskyblue",fg="white", font=("Times", 36,"bold"),bd=5,relief=GROOVE)

    	# create a Name label
        name = Label(root, text="NAME", bg="white", font=("Calibri", 26,"bold"))

    	# create a Course label
        course = Label(root, text="BRANCH", bg="white", font=("Calibri", 26,"bold"))

    	# create a Semester label
        sem = Label(root, text="SEMESTER", bg="white", font=("Calibri", 26,"bold"))

    	# create a Form No. lable
        form_no = Label(root, text="ROLL NO.", bg="white", font=("Calibri", 26,"bold"))

    	# create a Contact No. label
        contact_no = Label(root, text="CONTACT NO.", bg="white", font=("Calibri", 26,"bold"))

    	# create a Email id label
        email_id = Label(root, text="EMAIL ID", bg="white", font=("Calibri", 26,"bold"))

    	# create a address label
        address = Label(root, text="ADDRESS", bg="white", font=("Calibri", 26,"bold"))

        # create a username label
        username = Label(root, text="USERNAME", bg="white",font=("Calibri", 26,"bold"))

        # create a password label
        password = Label(root, text="PASSWORD", bg="white", font=("Calibri", 26,"bold"))

    	# grid method is used for placing
    	# the widgets at respective positions
    	# in table like structure .
        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        course.grid(row=2, column=0)
        sem.grid(row=3, column=0)
        form_no.grid(row=4, column=0)
        contact_no.grid(row=5, column=0)
        email_id.grid(row=6, column=0)
        address.grid(row=7, column=0)
        username.grid(row=8, column=0)
        password.grid(row=9, column=0)


    	# create a text entry box
    	# for typing the information
        name_field = Entry(root)
        course_field = Entry(root)
        sem_field = Entry(root)
        form_no_field = Entry(root)
        contact_no_field = Entry(root)
        email_id_field = Entry(root)
        address_field = Entry(root)
        username_field = Entry(root)
        password_field = Entry(root)

    	# bind method of widget is used for
    	# the binding the function with the events

    	# whenever the enter key is pressed
    	# then call the focus1 function
        name_field.bind("<Return>", focus1)

    	# whenever the enter key is pressed
    	# then call the focus2 function
        course_field.bind("<Return>", focus2)

    	# whenever the enter key is pressed
    	# then call the focus3 function
        sem_field.bind("<Return>", focus3)

    	# whenever the enter key is pressed
    	# then call the focus4 function
        form_no_field.bind("<Return>", focus4)

    	# whenever the enter key is pressed
    	# then call the focus5 function
        contact_no_field.bind("<Return>", focus5)

    	# whenever the enter key is pressed
    	# then call the focus6 function
        email_id_field.bind("<Return>", focus6)

        # whenever the enter key is pressed
    	# then call the focus6 function
        address_field.bind("<Return>", focus7)

        # whenever the enter key is pressed
    	# then call the focus6 function
        username_field.bind("<Return>", focus8)

    	# grid method is used for placing
    	# the widgets at respective positions
    	# in table like structure .
        name_field.grid(row=1, column=1, ipadx="200")
        course_field.grid(row=2, column=1, ipadx="200")
        sem_field.grid(row=3, column=1, ipadx="200")
        form_no_field.grid(row=4, column=1, ipadx="200")
        contact_no_field.grid(row=5, column=1, ipadx="200")
        email_id_field.grid(row=6, column=1, ipadx="200")
        address_field.grid(row=7, column=1, ipadx="200")
        username_field.grid(row=8, column=1, ipadx="200")
        password_field.grid(row=9, column=1, ipadx="200")

    	# call excel function
        excel()

    	# create a Submit Button and place into the root window
        submit = Button(root, text="REGISTER",bg="gold", fg="Black",
    							activebackground="red", command=insert,font=("Calibri", 36))
        submit.grid(row=10, column=1)



def login():
    global login_screen
    login_screen = Toplevel(main_screen)
    login_screen.title("Login")
    login_screen.geometry("900x700")
    login_screen.configure(background="white")
    login_screen.bg_img=PhotoImage(file="C:/Users/HP/Desktop/FRAS/FRAS/cp.png")
    Label(login_screen, text="Please enter details below to login",bg="deepskyblue",fg="white", width="300",font=("Times", 36,"bold"),bd=5,relief=GROOVE).pack()
    Label(login_screen, text="",bg="white").pack()
    bg_lbl=Label(login_screen,image=login_screen.bg_img).pack()

    global username_verify
    global password_verify

    username_verify = StringVar()
    password_verify = StringVar()

    global username_login_entry
    global password_login_entry

    Label(login_screen, text="Username * ",bg="white",font=("Calibri", 28,"bold")).pack()
    username_login_entry = Entry(login_screen, textvariable=username_verify ,font=("Times", 16,"bold"))

    username_login_entry.pack()
    Label(login_screen, text="",bg="white").pack()
    Label(login_screen, text="Password * ",bg="white",font=("Calibri", 28,"bold")).pack()
    password_login_entry = Entry(login_screen, textvariable=password_verify, show= '*',font=("Calibri", 16,"bold"))
    password_login_entry.pack()
    Label(login_screen, text="",bg="white").pack()



    Button(login_screen, activebackground="Red",text="Login",bg="gold", width=16, height=2,command=validate,font=("Times", 30,"bold")).pack()

def validate():

    if  username_login_entry.get() == "":
        root = Tk()
        w = Label(root, text='Please enter username to proceed')
        w.pack()
        root.mainloop()
    elif password_login_entry.get() == "" :
        root = Tk()
        w = Label(root, text='Please enter password to proceed')
        w.pack()
        root.mainloop()
    else:
        status = match()
        if(status):
            root = Tk()
            w = Label(root, text='Login Success')
            w.pack()
            root.mainloop()
            menu()
        else:
            root = Tk()
            w = Label(root, text='Invalid')
            w.pack()
            root.mainloop()


def match():
    if username_login_entry.get() ==  "admin" and password_login_entry.get() == "123" :
        return True
    else :
        return False

def menu():
    global m_screen
    m_screen = Tk()
    m_screen.geometry("2000x2000")
    m_screen.configure(background="white")
    m_screen.title("Account Login")
    Label(text="Choose from the buttons below", bg="deepskyblue",fg="white", width="300", height="2", font=("Times", 36,"bold"),bd=5,relief=GROOVE).pack()
    Label(text="",bg="white").pack()
    Button(text="Check Camera", height="2", width="20",bg="gold",command=checkCamera,activebackground="red",font=("Calibri",20)).pack()
    Label(text="",bg="white").pack()
    Button(text="Capture Faces", height="2", width="20",bg="gold",command=CaptureFaces,activebackground="red",font=("Calibri",20)).pack()
    Label(text="",bg="white").pack()
    Button(text="Train Image", height="2", width="20",bg="gold",command=Trainimages,activebackground="red",font=("Calibri",20)).pack()
    Label(text="",bg="white").pack()
    Button(text="Recognize and Attendance", height="2", width="20",bg="gold",command=RecognizeFaces,activebackground="red",font=("Calibri",20)).pack()
    m_screen.mainloop()


def checkCamera():
        check_camera.camer()
    # --------------------------------------------------------------
    # calling the take image function form capture image.py file

def CaptureFaces():
        Capture_Image.takeinput()
    # -----------------------------------------------------------------
    # calling the train images from train_images.py file

def Trainimages():
        Train_Image.TrainImages()
    # --------------------------------------------------------------------
    # calling the recognize_attendance from recognize.py file

def RecognizeFaces():
        Recognize.recognize_attendence()

def main_account_screen():
    global main_screen
    main_screen = Tk()
    main_screen.geometry("1000x650")
    main_screen.configure(background="white")
    main_screen.title("Account Login")
    main_screen.logo_icon=PhotoImage(file="C:/Users/HP/Desktop/FRAS/FRAS/lock02.png")
    #main_screen.bg_img=PhotoImage(file="C:/Users/Asus/Desktop/time.png")
    Label(text="Automated Real Time Face Attendance System", bg="deepskyblue",fg="white", width="300", height="2", font=("Times", 36,"bold"),bd=5,relief=GROOVE).pack()
    #bg_lbl=Label(main_screen,image=main_screen.bg_img).pack()
    Label(text="",bg="white").pack()
    LogoLabel=Label(main_screen,image=main_screen.logo_icon,width="220",height="250",bg="white").pack()
    Button(text="Login",bg="gold", height="2", width="15", command=login,activebackground="red",font=("Arial",25,"bold")).pack()
    Label(text="",bg="white").pack()
    Button(text="Register",bg="gold", height="2", width="15", command=register,activebackground="red",font=("Arial",25,"bold")).pack()

    main_screen.mainloop()


main_account_screen()
